from PySide6.QtWidgets import (QMenu, QMainWindow, QVBoxLayout,
                               QHBoxLayout, QWidget, QMenuBar, QCheckBox,
                               QSlider, QLabel, QFileDialog, QMessageBox,
                               QDialog, QTextEdit, QPushButton, QGridLayout, 
                               QTabWidget, QScrollArea, QSpinBox, QToolButton)

from core import (DataPool, StatisticsEngine, DatasetEntry, CowData, DataModel, 
                  GraphEngine, ALL_METRICS)
import sys
import os
import traceback
import pyqtgraph as pg
import numpy as np
from PySide6.QtGui import QAction, QIcon
from PySide6.QtCore import Qt, QTimer
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from dataset import VendorExcelParser, DatasetManager

home_path = os.path.expanduser('~')
log_file_path = os.path.join(home_path , 'RTT_logs')
base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
icon_file_path = os.path.join(base_path,'SJIT_icon.ico')

CSV_AUTO_CONVERT = False

# Default visibility settings for plots when opening a log file
DEFAULT_PLOT_VISIBILITY = {
    'cow_current_temp': True,
    'cow_station_temp': True,
    'cow_avg_temp': True,
    'cow_current_activity': True,
    'cow_station_activity': True,
    'cow_avg_activity': True
}

class TimeAxisItem(pg.AxisItem):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.start_time = None
        self.mode = "relative"  # relative | daily

    def set_start_time(self, start_time):
        self.start_time = start_time
    
    def set_daily_mode(self, enabled=True):
        if enabled:
            self.mode = "daily"
            self.start_time = None
        else:
            self.mode = "relative"

    def tickValues(self, minVal, maxVal, size):
        if self.mode == "daily":
            spacing = 3600  # 60 minutes in seconds

            # 0~86400 범위 고정
            ticks = []
            start = int(minVal // spacing) * spacing
            end = int(maxVal // spacing + 1) * spacing

            values = []
            for v in range(start, end + spacing, spacing):
                if 0 <= v <= 86400:
                    values.append(v)

            ticks.append((spacing, values))
            return ticks

        return super().tickValues(minVal, maxVal, size)

    def tickStrings(self, values, scale, spacing):
        # 하루 0~24시간 고정 모드
        if self.mode == "daily":
            strings = []
            for v in values:
                if v < 0 or v > 86400:
                    strings.append("")
                    continue
                h = int(v // 3600)
                m = int((v % 3600) // 60)
                strings.append(f"{h:02d}:{m:02d}")
            return strings

        # if self.start_time is None:
        return super().tickStrings(values, scale, spacing)


class CustomViewBox(pg.ViewBox):
    MAX_X_RANGE = 3600  # Maximum X-axis range in seconds (1 hour) - default value

    def __init__(self, *args, analyzer=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.analyzer = analyzer  # Reference to GasAnalyzer for dynamic max_x_range

    def mouseDragEvent(self, ev, axis=None):
        # Override to only allow X-axis panning in the plot area
        if ev.button() == Qt.MouseButton.LeftButton:
            # Check if drag started in plot area
            pos = ev.buttonDownPos()
            vb_rect = self.sceneBoundingRect()

            if vb_rect.contains(pos):
                # In plot area: only allow X-axis panning
                ev.accept()

                # Set flag to prevent Y auto range from being disabled during X-only pan
                if self.analyzer:
                    self.analyzer._updating_ranges = True

                # Store current Y range before panning
                current_y_range = self.viewRange()[1]

                # Let pyqtgraph handle the pan
                super().mouseDragEvent(ev, axis)

                # Restore Y range after panning
                self.setYRange(current_y_range[0], current_y_range[1], padding=0)

                # Enforce maximum X-axis range
                if ev.isFinish():
                    self._enforce_max_x_range()
                    # Disable X auto range and real time view after X-axis drag
                    if self.analyzer:
                        if self.analyzer.x_auto_range_action.isChecked():
                            self.analyzer.x_auto_range_action.setChecked(False)
                        if self.analyzer.realtime_view_action.isChecked():
                            self.analyzer.realtime_view_action.setChecked(False)
                        # Update stored X range
                        self.analyzer.stored_x_range = self.viewRange()[0]
                        from PySide6.QtCore import QTimer
                        QTimer.singleShot(50, lambda: setattr(self.analyzer, '_updating_ranges', False))
            else:
                # On axes: allow normal behavior
                super().mouseDragEvent(ev, axis)
        else:
            # For other mouse buttons, use default behavior
            super().mouseDragEvent(ev, axis)

    def wheelEvent(self, ev, axis=None):
        # Get the position of the mouse event
        pos = ev.pos()
        vb_rect = self.sceneBoundingRect()

        # Check if mouse is in the plot area (not on axes)
        if vb_rect.contains(pos):
            delta = ev.delta()

            # In plot area: only allow X-axis zoom/pan
            if delta > 0:
                self.scaleBy(x=0.9, y=1.0)
            else:
                self.scaleBy(x=1.1, y=1.0)

            # Enforce maximum X-axis range
            self._enforce_max_x_range()

            # Disable X auto range and real time view after X-axis zoom
            if self.analyzer:
                if self.analyzer.x_auto_range_action.isChecked():
                    self.analyzer.x_auto_range_action.setChecked(False)
                if self.analyzer.realtime_view_action.isChecked():
                    self.analyzer.realtime_view_action.setChecked(False)
                # Update stored X range
                self.analyzer.stored_x_range = self.viewRange()[0]

            ev.accept()
        else:
            # On axes: allow normal behavior
            super().wheelEvent(ev, axis)

    def _enforce_max_x_range(self):
        print("MAX RANGE ENFORCED")
        """Limit the X-axis visible range to max_x_range seconds"""
        # Get max range from analyzer if available, otherwise use default
        max_range = self.analyzer.max_x_range if self.analyzer else self.MAX_X_RANGE

        view_range = self.viewRange()
        x_min, x_max = view_range[0]
        current_range = x_max - x_min

        if current_range > max_range:
            # Calculate the center of the current view
            center = (x_min + x_max) / 2
            # Set new range centered on the current view
            new_x_min = center - max_range / 2
            new_x_max = center + max_range / 2
            self.setXRange(new_x_min, new_x_max, padding=0)


class ErrorDialog(QDialog):
    def __init__(self, title, message, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(600, 400)

        layout = QVBoxLayout(self)

        # Text edit for error message (read-only but copyable)
        self.text_edit = QTextEdit()
        self.text_edit.setPlainText(message)
        self.text_edit.setReadOnly(True)
        layout.addWidget(self.text_edit)

        # OK button
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        layout.addWidget(ok_button)


class SettingsDialog(QDialog):
    def __init__(self, current_max_x_range, parent=None):
        super().__init__(parent)
        self.setWindowTitle("X-Axis Range Settings")
        self.setModal(True)
        self.resize(400, 150)

        layout = QVBoxLayout(self)

        # Description label
        desc_label = QLabel("Set the maximum time range displayed on the X-axis:")
        layout.addWidget(desc_label)

        # Input section
        input_layout = QHBoxLayout()

        # Spinbox for time value
        self.time_spinbox = QSpinBox()
        self.time_spinbox.setMinimum(1)
        self.time_spinbox.setMaximum(24)  # Max 24 hours
        self.time_spinbox.setValue(current_max_x_range // 3600)  # Convert seconds to hours
        self.time_spinbox.setSuffix(" hour(s)")
        input_layout.addWidget(QLabel("Max X Range:"))
        input_layout.addWidget(self.time_spinbox)
        input_layout.addStretch()

        layout.addLayout(input_layout)

        # Info label
        info_label = QLabel("Note: This setting limits the maximum visible time range when using X auto scale.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(info_label)

        layout.addStretch()

        # Buttons
        button_layout = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)

    def get_max_x_range(self):
        """Returns the max X range in seconds"""
        return self.time_spinbox.value() * 3600  # Convert hours to seconds

class CowAnalyzer(QMainWindow):
    def __init__(self): # 클래스가 처음 만들어질 때, 호출
        super().__init__()

        self.data_pool = DataPool()
        self.data_model = DataModel()

        self.global_temp_range = None
        self.global_act_range = None

        self.start_time = None  # Store the start time for time axis formatting
        self.graph_widget = None

        # open file name
        self.file_name = None

        # Analysis lines
        self.line_a = None
        self.line_b = None

        # Real-time monitoring
        self.log_monitor_timer = None
        self.last_file_position = 0
        self.current_file_path = None

        # Store current view ranges
        self.stored_x_range = None
        self.stored_y_ranges = {}

        # Flag to prevent recursive range updates
        self._updating_ranges = False
        # Flag to indicate legend is being toggled (to prevent auto range disable)
        self._toggling_legend = False

        self.hover_proxy = None
        self.v_line = None
        self.h_line = None
        self.hover_label = None
        self.hover_curves = []
        self._hover_updating = False

        # Settings
        self.max_x_range = 86400     # Default: 24 hour in seconds

        self.clear_action = QAction("Clear all datasets", self)
        self.clear_action.triggered.connect(self.clear_all_datasets)

        self.setup_ui()
        self.setup_connections()

        self.dataset_manager = DatasetManager()

    def _get_build_date(self):
        """Read build date from build_date.txt (generated by rebuild.cmd)"""
        try:
            # PyInstaller frozen executable
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            date_file = os.path.join(base_path, 'build_date.txt')
            with open(date_file, 'r') as f:
                return f.read().strip()
        except Exception:
            return 'dev'
    
    


    def setup_ui(self): # UI구성, 애플리케이션 외형 생성
        print("setup_ui()")
        self.setWindowTitle(f"Cow Ear Tag Data Analyzer (Build: {self._get_build_date()})")
        self.setGeometry(100, 100, 1200, 800)

        # Create menu bar
        menubar = self.menuBar()
        file_menu = menubar.addMenu('File')

        self.open_action = QAction('Open Log File', self)
        file_menu.addAction(self.open_action)

        self.export_excel_action = QAction('Export to Excel', self)
        file_menu.addAction(self.export_excel_action)

        # Settings menu
        settings_menu = menubar.addMenu('Settings')

        self.settings_action = QAction('X-Axis Range Settings...', self)
        settings_menu.addAction(self.settings_action)

        # Create toolbar for toggle buttons
        toolbar = self.addToolBar("View Controls")
        toolbar.setMovable(False)

        # X-axis auto range toggle
        self.x_auto_range_action = QAction("⮂ X Auto", self)
        self.x_auto_range_action.setCheckable(True)
        self.x_auto_range_action.setChecked(True)  # Default enabled
        toolbar.addAction(self.x_auto_range_action)

        # Y-axis auto range toggle
        self.y_auto_range_action = QAction("⮃ Y Auto", self)
        self.y_auto_range_action.setCheckable(True)
        self.y_auto_range_action.setChecked(True)  # Default enabled
        toolbar.addAction(self.y_auto_range_action)

        # Real time view toggle
        self.realtime_view_action = QAction("⇒ Real Time", self)
        self.realtime_view_action.setCheckable(True)
        self.realtime_view_action.setChecked(True)  # Default disabled
        toolbar.addAction(self.realtime_view_action)

        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Create main horizontal layout
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # Create left control panel
        self.left_panel = QWidget()
        self.left_panel.setFixedWidth(300)
        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)

        # -----------------------------
        # 좌측 탭 위젯
        # -----------------------------
        self.left_tabs = QTabWidget()
        left_layout.addWidget(self.left_tabs)

        # ==================================================
        # Tab 1 — Data
        # ==================================================
        self.data_tab = QWidget()
        self.left_tabs.addTab(self.data_tab, "Data")
        data_layout = QVBoxLayout(self.data_tab)

        # 데이터 스크롤 영역
        self.dataset_scroll = QScrollArea()
        self.dataset_scroll.setWidgetResizable(True)

        self.dataset_container = QWidget()
        self.dataset_layout = QVBoxLayout(self.dataset_container)
        self.dataset_layout.setAlignment(Qt.AlignTop)

        self.dataset_scroll.setWidget(self.dataset_container)
        data_layout.addWidget(self.dataset_scroll)

        # ==================================================
        # Tab 2 — Analysis
        # ==================================================
        self.analysis_tab = QWidget()
        self.left_tabs.addTab(self.analysis_tab, "Analysis")
        analysis_layout = QVBoxLayout(self.analysis_tab)
        
        self.lines_checkbox = QCheckBox("Enable Lines A & B")
        analysis_layout.addWidget(self.lines_checkbox)

        self.analysis_text = QTextEdit()
        self.analysis_text.setReadOnly(True)
        analysis_layout.addWidget(self.analysis_text)

        # -----------------------------
        # Statistics Section
        # -----------------------------
        left_layout.addWidget(QLabel("Statistics (Visible Range)"))

        stats_widget = QWidget()
        stats_layout = QGridLayout(stats_widget)
        stats_layout.setSpacing(4)

        self.stats_time_a = QLabel("-")
        self.stats_time_b = QLabel("-")

        stats_layout.addWidget(QLabel("Time A"), 0, 0)
        stats_layout.addWidget(self.stats_time_a, 0, 1)

        stats_layout.addWidget(QLabel("Time B"), 1, 0)
        stats_layout.addWidget(self.stats_time_b, 1, 1)

        left_layout.addWidget(stats_widget)
        left_layout.addStretch()

        # ==================================================
        # GRAPH
        # ==================================================
        self.setup_graph()

        # ==================================================
        # Add to Main Layout (한 번만!)
        # ==================================================
        main_layout.addWidget(self.left_panel)
        main_layout.addWidget(self.graph_widget, 1)

        try:
            self.setWindowIcon(QIcon(icon_file_path))
        except:
            pass


    def setup_graph(self): # 그래프 설정(그래프 기능의 핵심)
        print("setup_graph()")
        # Create PlotWidget with custom ViewBox for better mouse control
        self.graph_widget = pg.PlotWidget(viewBox=CustomViewBox(analyzer=self), axisItems={'bottom': TimeAxisItem(orientation='bottom')})
        self.graph_widget.setLabel('bottom', 'Time')
        self.graph_widget.showGrid(x=True, y=True)
        self.legend = self.graph_widget.addLegend()
        # Connect legend click events using the scene's mouse click signal
        self.legend.scene().sigMouseClicked.connect(self.on_legend_clicked)

        # Create multiple Y-axis labels on the left side
        plot_item = self.graph_widget.getPlotItem()
        self.main_vb = plot_item.vb  # Store reference to main ViewBox

        # Activity ViewBox for Cow Data
        self.activity_vb = CustomViewBox(analyzer=self)

        # Left axis for Temperature
        plot_item.setLabel('left', 'Temperature (°C)', color='red')

        # Create additional axes
        # Activity Axis
        self.activity_axis = pg.AxisItem('right')
        self.activity_axis.setLabel('Activity', color='green')

        # Add the additional axes to the plot layout
        # Right side: Activity
        plot_item.layout.addItem(self.activity_axis, 2, 3)


        # Add ViewBoxes to scene
        plot_item.scene().addItem(self.activity_vb)


        # Link X-axis but keep Y-axis independent
        self.activity_vb.setXLink(self.main_vb)


        # Set X axis limits to prevent going below 0
        self.main_vb.setLimits(xMin=0)
        self.activity_vb.setLimits(xMin=0)


        # Configure mouse behavior: drag to pan, wheel to zoom X-axis only
        self.main_vb.setMouseMode(pg.ViewBox.PanMode)

        # Connect ViewBox range change signal to update statistics when X range changes
        self.main_vb.sigRangeChanged.connect(self.on_range_changed)

        # Connect all ViewBox range changes to detect manual Y-axis zoom/pan
        self.main_vb.sigRangeChanged.connect(lambda vb, ranges: self.on_view_range_changed(vb, ranges))
        self.activity_vb.sigRangeChanged.connect(lambda vb, ranges: self.on_view_range_changed(vb, ranges))

        # Initialize analysis lines as None (will be created when needed)
        self.line_a = None
        self.line_b = None

        # Store original data ranges for scaling
        self.c_avg_range = None
        self.c_delta_range = None
        self.temp_range = None
        self.humidity_range = None

        # -----------------------------
        # Crosshair cursor system
        # -----------------------------
        self.vline = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen(color='y'))
        self.main_vb.addItem(self.vline, ignoreBounds=True)

        # self.cursor_shadow = pg.TextItem(
        #     anchor=(0, 1),
        #     color=(0, 0, 0, 150) # 반투명 검정
        # )
        self.cursor_text = pg.TextItem(
            anchor=(0,1),
            color='w',
            fill=pg.mkBrush(0, 0, 0, 90), # 반투명 배경
            # border=pg.mkPen((255,255,255,150)) # 텍스트 상자
        )
        # self.cursor_shadow.setZValue(99)
        self.cursor_text.setZValue(100) # 그래프보다 위에 표시
        # self.main_vb.addItem(self.cursor_shadow, ignoreBounds=True)
        self.main_vb.addItem(self.cursor_text, ignoreBounds=True)

        self.mouse_proxy = pg.SignalProxy(
            self.graph_widget.scene().sigMouseMoved,
            rateLimit=60,
            slot=self.on_mouse_moved
        ) 

        # Data time reference
        self.start_time = None
        self.current_time_seconds = None
    

    def on_mouse_moved(self, evt):
        # print("HOVER EVENT")
        if getattr(self, "current_time_seconds", None) is None:
            # print("NO TIME DATA")
            return
        if self._hover_updating:
            print("NO HOVER")
            return
        self._hover_updating = True

        try:
            pos = evt[0]

            mouse_point = self.main_vb.mapSceneToView(pos)
            x = mouse_point.x()
            y = mouse_point.y()

            # 커서선 이동
            self.vline.setPos(x)

            time_array = np.array(self.current_time_seconds)
            if len(time_array) == 0:
                print("time_array is 0")
                return
            # 가장 가까운 데이터 index
            idx = np.argmin(np.abs(time_array - x))

            data_time = self.start_time + timedelta(seconds=time_array[idx])
            time_str = data_time.strftime("%H:%M")
            text_lines = [time_str]

            # 모든 visible dataset 검사
            for entry in self.get_visible_entries():
                for metric, curve in entry.curves.items():
                    if not curve.isVisible():
                        continue

                    xdata, ydata = curve.getData()
                    if ydata is None or idx >= len(ydata):
                        continue

                    curve_y = ydata[idx]
                    text_lines.append(f"{metric}: {curve_y:.2f}")

            text = "\n".join(text_lines)

            offset = 0.02*(self.main_vb.viewRange()[0][1]-self.main_vb.viewRange()[0][0])
            self.cursor_text.setText(text)
            # self.cursor_shadow.setText(text)

            # self.cursor_shadow.setPos(x + offset + 0.1, y - 0.1) # 그림자는 살짝 아래쪽
            self.cursor_text.setPos(x + offset, y)
        except Exception as e:
            print("HOVER ERROR:", e)
            traceback.print_exc()
        finally:
            self._hover_updating = False


    def on_legend_clicked(self, event):
        """Handles clicks on the legend to toggle plot visibility."""
        # Set flag to prevent auto range from being disabled during legend toggle
        self._toggling_legend = True

        # The legend items are a list of (sample, label) tuples
        for sample, label in self.legend.items:
            # Check if the click was on this legend item (sample or label area)
            sample_rect = sample.sceneBoundingRect()
            label_rect = label.sceneBoundingRect()
            click_pos = event.scenePos()

            if sample_rect.contains(click_pos) or label_rect.contains(click_pos):
                # Determine which actual plot corresponds to the clicked label
                if label.text == 'Current Temp':
                    # Toggle visibility
                    if self.cow_curr_temp_curve:
                        # Toggle visibility flag
                        self.cow_current_temp_visible = not self.cow_current_temp_visible

                        # Check if there's actual data
                        is_visible = self.cow_current_temp_visible

                        self.cow_curr_temp_curve.setVisible(is_visible)

                        # Update legend item visibility (like other graphs do)
                        if hasattr(sample, 'item'):
                            sample.item.setVisible(self.cow_current_temp_visible)

                        # Force legend and graph update to reflect changes immediately
                        if hasattr(self, 'legend'):
                            self.legend.update()
                        self.graph_widget.update()

                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()

                
                elif label.text == 'Station Temp':
                    if self.cow_stn_temp_curve:
                        # Toggle visibility flag
                        self.cow_station_temp_visible = not self.cow_station_temp_visible

                        # Check if there's actual data
                        is_visible = self.cow_station_temp_visible

                        self.cow_stn_temp_curve.setVisible(is_visible)

                        # Update legend item visibility (like other graphs do)
                        if hasattr(sample, 'item'):
                            sample.item.setVisible(self.cow_station_temp_visible)

                        # Force legend and graph update to reflect changes immediately
                        if hasattr(self, 'legend'):
                            self.legend.update()
                        self.graph_widget.update()

                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()

                
                elif label.text == 'Avg Temp':
                    if self.cow_avg_temp_curve:
                        # Toggle visibility flag
                        self.cow_avg_temp_visible = not self.cow_avg_temp_visible

                        # Check if there's actual data
                        is_visible = self.cow_avg_temp_visible

                        self.cow_avg_temp_curve.setVisible(is_visible)

                        if hasattr(sample, 'item'):
                            sample.item.setVisible(is_visible)

                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()

                
                elif label.text == 'Current Activity':
                    if self.cow_curr_act_curve:
                        # Toggle visibility flag
                        self.cow_current_activity_visible = not self.cow_current_activity_visible

                        # Check if there's actual data
                        is_visible = self.cow_current_activity_visible

                        self.cow_curr_act_curve.setVisible(is_visible)

                        # Toggle axis visibility
                        if hasattr(self, 'activity_axis'):
                            self.activity_axis.setVisible(is_visible)

                        if hasattr(sample, 'item'):
                            sample.item.setVisible(is_visible)

                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()

                
                elif label.text == 'Station Activity':
                    if self.cow_stn_act_curve:
                        is_visible = not self.cow_stn_act_curve.isVisible()
                        self.cow_station_activity_visible = is_visible  # Sync state
                        
                        
                        if hasattr(sample, 'item'):
                            sample.item.setVisible(is_visible)
                        
                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()
                        self.cow_stn_act_curve.setVisible(is_visible)

                
                elif label.text == 'Avg Activity':
                    if self.cow_avg_act_curve:
                        is_visible = not self.cow_avg_act_curve.isVisible()
                        self.cow_avg_activity_visible = is_visible  # Sync state
                        
                        
                        if hasattr(sample, 'item'):
                            sample.item.setVisible(is_visible)
                        
                        # Set flag before update_axis_layout to prevent auto range disable
                        self._updating_ranges = True
                        self.update_axis_layout()
                        self.cow_avg_act_curve.setVisible(is_visible)

                event.accept()
                # Reset legend toggle flag after a short delay to allow all updates to complete
                QTimer.singleShot(50, lambda: setattr(self, '_toggling_legend', False))
                return  # Exit after handling the first matching click

        # Reset flag if no legend item was clicked
        self._toggling_legend = False

    def on_x_auto_range_toggled(self, checked):
        """Handle X auto range toggle - re-apply auto range when enabled"""
        if not checked:
            return

        start_time, time_seconds = self.data_model.get_time_reference()
        if not time_seconds:
            return

        self._updating_ranges = True
        try:
            self.apply_x_auto_range()
        finally:
            QTimer.singleShot(
                50,
                lambda: setattr(self, '_updating_ranges', False)
            )

    def apply_x_auto_range(self):
        """Apply auto range to X axis based on current data"""
        start_time, time_seconds = self.data_model.get_time_reference()
        if not time_seconds:
            return

        max_time = max(time_seconds)

        if max_time > self.max_x_range:
            x_min = max_time - self.max_x_range
            x_max = max_time
        else:
            x_min = 0
            x_max = max_time

        self.main_vb.setXRange(x_min, x_max, padding=0.02)
        self.stored_x_range = (x_min, x_max)


    def on_y_auto_range_toggled(self, checked):
        """Handle Y auto range toggle - re-apply auto range when enabled"""
        if not checked:
            return

        start_time, time_seconds = self.data_model.get_time_reference()
        if not time_seconds:
            return

        self._updating_ranges = True
        try:
            self.apply_y_auto_ranges()
        finally:
            QTimer.singleShot(
                50,
                lambda: setattr(self, '_updating_ranges', False)
            )


    def apply_y_auto_ranges(self):
        """Apply auto range to all Y axes based on current data"""
        entry = self.data_model.get_primary_entry()
        if not entry:
            return

        cow = entry.cow
        if not cow.timestamps:
            return

        cow_temps = [
            v for v in (
                cow.cow_current_temp +
                cow.cow_station_temp +
                cow.cow_avg_temp
            )
            if v is not None
        ]

        cow_acts = [
            v for v in (
                cow.cow_current_activity +
                cow.cow_station_activity +
                cow.cow_avg_activity
            )
            if v is not None
        ]

        if cow_temps:
            min_val, max_val = min(cow_temps), max(cow_temps)
            y_min, y_max = self.calculate_y_range_for_quarter(min_val, max_val)
            self.main_vb.setYRange(y_min, y_max)
            self.stored_y_ranges['temp'] = (y_min, y_max)
            self.graph_widget.setLabel(
                'left',
                f'Temperature (°C): {y_min:.1f} - {y_max:.1f}',
                color='red'
            )

        if cow_acts:
            min_val, max_val = min(cow_acts), max(cow_acts)
            y_min, y_max = self.calculate_y_range_for_quarter(min_val, max_val)
            self.activity_vb.setYRange(y_min, y_max)
            self.stored_y_ranges['activity'] = (y_min, y_max)
            self.activity_axis.setLabel(
                f'Activity: {int(y_min)} - {int(y_max)}',
                color='green'
            )


    def on_view_range_changed(self, view_box, ranges):
        """Store current X and Y ranges when user manually zooms/pans"""
        if getattr(self, "_hover_updating", False):
            return
        
        # Disable flag to prevent recursive updates
        if not hasattr(self, '_updating_ranges'):
            self._updating_ranges = False

        if self._updating_ranges:
            return

        # Don't disable auto range if we're just toggling legend visibility
        if not hasattr(self, '_toggling_legend'):
            self._toggling_legend = False
        if self._toggling_legend:
            return

        # Determine which ViewBox triggered the change
        vb_name = None
        if view_box == self.main_vb:
            vb_name = 'temp'
        elif hasattr(self, 'activity_vb') and view_box == self.activity_vb:
            vb_name = 'activity'

        print(f"[DEBUG] on_view_range_changed called from ViewBox: {vb_name}")
        print("MAIN RANGE:", self.main_vb.viewRange())
        if hasattr(self, "activity_vb"):
            print("ACT RANGE:", self.activity_vb.viewRange())

        # Get current ranges
        x_range = ranges[0]
        y_range = ranges[1]

        # For main_vb, check Y first (to prioritize Y-axis drag detection)
        # For other ViewBoxes, only check Y
        if vb_name:
            if vb_name in self.stored_y_ranges:
                old_y_range = self.stored_y_ranges[vb_name]
                old_y_span = old_y_range[1] - old_y_range[0]
                y_span = y_range[1] - y_range[0]

                # Calculate percentage change (use 0.1% as threshold for meaningful change)
                y_min_change_pct = abs(y_range[0] - old_y_range[0]) / max(old_y_span, 0.001) * 100
                y_max_change_pct = abs(y_range[1] - old_y_range[1]) / max(old_y_span, 0.001) * 100
                y_span_change_pct = abs(y_span - old_y_span) / max(old_y_span, 0.001) * 100

                # Y changed if position changed OR span changed significantly
                y_changed = (y_min_change_pct > 0.1 or y_max_change_pct > 0.1 or y_span_change_pct > 0.1)

                if y_changed:
                    # Y range changed for this ViewBox
                    print(f"[DEBUG] Y changed for {vb_name}: {old_y_range} -> {y_range} (min: {y_min_change_pct:.2f}%, max: {y_max_change_pct:.2f}%, span: {y_span_change_pct:.2f}%)")
                    self.stored_y_ranges[vb_name] = y_range

                    # Check if X also changed (indicating X drag, not Y drag)
                    # All ViewBoxes share X axis, so check main_vb's X range
                    x_also_changed = False
                    if self.stored_x_range is not None and hasattr(self, 'main_vb'):
                        current_x = self.main_vb.viewRange()[0]
                        old_x_range = self.stored_x_range
                        old_x_span = old_x_range[1] - old_x_range[0]
                        x_span_temp = current_x[1] - current_x[0]
                        x_min_pct = abs(current_x[0] - old_x_range[0]) / max(old_x_span, 0.001) * 100
                        x_max_pct = abs(current_x[1] - old_x_range[1]) / max(old_x_span, 0.001) * 100
                        x_span_pct = abs(x_span_temp - old_x_span) / max(old_x_span, 0.001) * 100
                        x_also_changed = (x_min_pct > 0.1 or x_max_pct > 0.1 or x_span_pct > 0.1)

                    if x_also_changed:
                        # Y change is side effect of X drag - don't disable Y auto range
                        print(f"[DEBUG] Y change on {vb_name} is due to X drag - ignoring Y change")
                        # Update X range to prevent false X change detection on next call
                        # (Always update, not just for main_vb)
                        self.stored_x_range = current_x
                        return
                    else:
                        # Pure Y drag - disable Y auto range ONLY
                        # Y drag shouldn't affect X range, but check if X changed slightly due to geometry
                        # and update stored X to prevent false X detection on next event
                        current_x_check = self.main_vb.viewRange()[0]
                        if self.stored_x_range is not None:
                            # Update X range if it changed slightly (but not enough to trigger x_also_changed)
                            self.stored_x_range = current_x_check

                        if self.y_auto_range_action.isChecked():
                            print(f"[DEBUG] Disabling Y auto range due to Y change on {vb_name}")
                            self.y_auto_range_action.setChecked(False)

                        # Don't check X if Y changed (Y drag takes priority)
                        return
            else:
                # First time storing - not a user change
                self.stored_y_ranges[vb_name] = y_range
                print(f"[DEBUG] First Y store for {vb_name}: {y_range}")

        # Only process X range changes from main_vb (to avoid duplicate processing from linked ViewBoxes)
        # Only reach here if Y didn't change
        if view_box == self.main_vb:
            # Check X range change - use percentage threshold for meaningful changes
            if self.stored_x_range is not None:
                old_x_range = self.stored_x_range
                old_x_span = old_x_range[1] - old_x_range[0]
                x_span = x_range[1] - x_range[0]

                # Calculate percentage change (use 0.1% as threshold for meaningful change)
                x_min_change_pct = abs(x_range[0] - old_x_range[0]) / max(old_x_span, 0.001) * 100
                x_max_change_pct = abs(x_range[1] - old_x_range[1]) / max(old_x_span, 0.001) * 100
                x_span_change_pct = abs(x_span - old_x_span) / max(old_x_span, 0.001) * 100

                # X changed if position changed OR span changed significantly
                x_changed = (x_min_change_pct > 0.1 or x_max_change_pct > 0.1 or x_span_change_pct > 0.1)

                if x_changed:
                    # X range changed - disable X auto range and real time ONLY
                    print(f"[DEBUG] X changed on main_vb: {old_x_range} -> {x_range} (min: {x_min_change_pct:.2f}%, max: {x_max_change_pct:.2f}%, span: {x_span_change_pct:.2f}%)")
                    self.stored_x_range = x_range

                    # Update all ViewBox Y ranges to current values to prevent false Y change detection
                    # (X axis drag causes ViewBox geometry recalculation which changes Y ranges)
                    self.stored_y_ranges['temp'] = self.main_vb.viewRange()[1]
                    if hasattr(self, 'activity_vb'):
                        self.stored_y_ranges['activity'] = self.activity_vb.viewRange()[1]

                    if self.x_auto_range_action.isChecked():
                        print(f"[DEBUG] Disabling X auto range due to X change")
                        self.x_auto_range_action.setChecked(False)
                    if self.realtime_view_action.isChecked():
                        print(f"[DEBUG] Disabling real time due to X change")
                        self.realtime_view_action.setChecked(False)
                    return
            else:
                # First time storing X range - not a user change
                self.stored_x_range = x_range
                print(f"[DEBUG] First X store: {x_range}")

    def calculate_y_range_for_quarter(self, min_val, max_val):
        """
        Calculate Y-axis range so that data fills 1/4 ~ 3/4 of the axis.

        If data range is [min_val, max_val], we want:
        - min_val to appear at 1/4 of Y-axis
        - max_val to appear at 3/4 of Y-axis

        So: y_min + (y_max - y_min) * 0.25 = min_val
            y_min + (y_max - y_min) * 0.75 = max_val

        Solving: data_range = max_val - min_val
                 data_range = (y_max - y_min) * 0.5
                 y_range = data_range / 0.5 = data_range * 2
                 y_min = min_val - data_range * 0.5
                 y_max = max_val + data_range * 0.5
        """
        if min_val == max_val:
            # Handle single value case
            margin = max(abs(min_val) * 0.1, 1.0)  # 10% margin or at least 1
            return max(0, min_val - margin * 2), max_val + margin * 2

        data_range = max_val - min_val
        y_min = min_val - data_range * 0.5
        y_max = max_val + data_range * 0.5

        # Ensure y_min is not negative for data that should start from 0
        # (This can be adjusted per data type if needed)
        y_min = max(0, y_min)

        return y_min, y_max

    def setup_connections(self):
        print("setup_connections()")
        self.open_action.triggered.connect(self.open_log_file)
        self.export_excel_action.triggered.connect(self.export_to_excel_menu)
        self.settings_action.triggered.connect(self.open_settings_dialog)
        self.lines_checkbox.toggled.connect(self.toggle_analysis_lines)

        # Connect auto range toggle actions
        self.x_auto_range_action.toggled.connect(self.on_x_auto_range_toggled)
        self.y_auto_range_action.toggled.connect(self.on_y_auto_range_toggled)


    def open_settings_dialog(self):
        """Open the settings dialog to configure max X range"""
        dialog = SettingsDialog(self.max_x_range, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Update max X range setting
            self.max_x_range = dialog.get_max_x_range()

            # If X auto range is enabled, reapply the range with new setting
            start_time, time_seconds = self.data_model.get_time_reference()
            if self.x_auto_range_action.isChecked() and time_seconds:
                self.apply_x_auto_range()
    
    def _calculate_global_y_ranges(self):
        """로드된 모든 데이터 기준 Y축 범위 1회 계산"""
        all_cows = [entry.cow for entry in self.data_model.get_all_entries()]

        all_temps, all_acts = GraphEngine.calculate_ranges(all_cows)

        if all_temps:
            self.global_temp_range = self.calculate_y_range_for_quarter(
                min(all_temps), max(all_temps)
            )

        if all_acts:
            self.global_act_range = self.calculate_y_range_for_quarter(
                min(all_acts), max(all_acts)
            )

    def open_log_file(self, file_name = False):
        if file_name == False:
            # Create a custom file dialog with sorting by modification time
            dialog = QFileDialog(self, "Open Log File", log_file_path)
            dialog.setNameFilter("Log files (*.log *.txt *.csv *.xlsx);;All files (*)")
            dialog.setFileMode(QFileDialog.FileMode.ExistingFile)

            # Set view mode to Detail to enable sorting
            dialog.setViewMode(QFileDialog.ViewMode.Detail)

            # Access the internal list view and set sorting
            from PySide6.QtCore import Qt
            from PySide6.QtWidgets import QListView, QTreeView

            # Find the tree view (detail view) and enable sorting by modification time
            for child in dialog.findChildren(QTreeView):
                child.sortByColumn(3, Qt.SortOrder.DescendingOrder)  # Column 3 is "Date Modified"
                child.setSortingEnabled(True)

            if dialog.exec() == QFileDialog.DialogCode.Accepted:
                file_path = dialog.selectedFiles()[0]
            else:
                file_path = None
        else:
            file_path = file_name
        
        if not file_path:
            return

        try:
            # Stop any existing monitoring timer first
            if self.log_monitor_timer:
                self.log_monitor_timer.stop()
                print("[DEBUG] open_log_file: Stopped existing monitoring timer")

            self.file_name = file_path
            self.current_file_path = file_path

            # Reset start_time to force new date/time baseline
            self.start_time = None

            # Clear analysis lines
            self.line_a = None
            self.line_b = None

            # Clear stored ranges before parsing
            self.stored_x_range = None
            self.stored_y_ranges = {}
            print("[DEBUG] open_log_file: Cleared stored ranges")

            # Reset CSV/Excel tracking indices
            self.last_csv_index = 0
            self.last_excel_index = 0

            # Reset plot visibility to default settings
            self.cow_current_temp_visible = DEFAULT_PLOT_VISIBILITY['cow_current_temp']
            self.cow_station_temp_visible = DEFAULT_PLOT_VISIBILITY['cow_station_temp']
            self.cow_avg_temp_visible = DEFAULT_PLOT_VISIBILITY['cow_avg_temp']
            self.cow_current_activity_visible = DEFAULT_PLOT_VISIBILITY['cow_current_activity']
            self.cow_station_activity_visible = DEFAULT_PLOT_VISIBILITY['cow_station_activity']
            self.cow_avg_activity_visible = DEFAULT_PLOT_VISIBILITY['cow_avg_activity']

            # Hide all graph axes before parsing (in case new file has no data for some graphs)
            if hasattr(self, 'activity_axis'):
                self.activity_axis.setStyle(showValues=False)

            # Prevent range change signals from disabling auto range during update
            print("[DEBUG] open_log_file: About to parse and update graph")
            self._updating_ranges = True
            self.parse_log_file(file_path)

            try:
                if file_path.endswith((".xlsx", ".xls")):
                    parser = VendorExcelParser()
                    dataset = parser.parse(file_path)

                    # store in new manager
                    self.dataset_manager.add(dataset)

                    print("[STEP1] Dataset loaded:", dataset.cow_id, dataset.available_metrics())

            except Exception as e:
                print("[STEP1] Dataset parsing skipped:", e)

            for cow in self.data_pool.all():
                key = cow.key
                entry = self.data_model.get_entry(key)
                if entry and entry.checkbox is None:
                    self.create_dataset_checkbox(key)
            self._calculate_global_y_ranges()
                    
            self.update_graph()
            print("[DEBUG] open_log_file: Graph updated")

            # Update all stored ranges to current values after graph update
            # This prevents subsequent range change signals from being detected as user changes
            self.stored_y_ranges['temp'] = self.main_vb.viewRange()[1]
            if hasattr(self, 'activity_vb'):
                self.stored_y_ranges['activity'] = self.activity_vb.viewRange()[1]

            # Enable auto range AFTER updating stored ranges
            print("[DEBUG] open_log_file: Enabling auto ranges...")
            self.x_auto_range_action.setChecked(True)
            print(f"[DEBUG] X auto range set: {self.x_auto_range_action.isChecked()}")
            self.y_auto_range_action.setChecked(True)
            print(f"[DEBUG] Y auto range set: {self.y_auto_range_action.isChecked()}")
            self.realtime_view_action.setChecked(True)
            print(f"[DEBUG] Real time set: {self.realtime_view_action.isChecked()}")

            # Delay re-enabling range change detection to ensure all auto ranges stay enabled
            QTimer.singleShot(200, lambda: setattr(self, '_updating_ranges', False))
            print("[DEBUG] open_log_file: Scheduled _updating_ranges = False after 200ms")


            # Get initial file position
            if '.csv' in file_path or '.xlsx' in file_path:
                QMessageBox.information(self, "Success", "Log file loaded successfully!")
            else:
                # Export to CSV/Excel only if CSV_AUTO_CONVERT is True
                if CSV_AUTO_CONVERT and self.data_pool:
                    dc = self.data_pool[list(self.data_pool.keys())[-1]].data

                    self.export_to_csv(file_path)
                    self.export_to_excel(file_path)
                    # Set indices after initial export
                    self.last_csv_index = len(dc.timestamps)
                    self.last_excel_index = len(dc.timestamps)
                    QMessageBox.information(self, "Success", "Log file loaded and CSV exported successfully!")
                else:
                    QMessageBox.information(self, "Success", "Log file loaded successfully!")

        except Exception as e:
            import traceback
            error_msg = f"Failed to process log file:\n\n{str(e)}\n\nFull traceback:\n{traceback.format_exc()}"
            error_dialog = ErrorDialog("Error", error_msg, self)
            error_dialog.exec()

    def on_dataset_toggled(self, key, state):
        print("TOGGLED!", key, state)

        entry = self.data_model.get_entry(key)
        if not entry:
            return

        entry.visible = bool(state)
        print("on_dataset_toggled : visible =", entry.visible)

        for cb in entry.metric_checkboxes:
            cb.blockSignals(True)
            cb.setChecked(entry.visible)
            cb.blockSignals(False)

        for metric_key in entry.metric_visible:
            entry.set_metric_visible(metric_key, entry.visible)

        self.update_graph()

    def create_dataset_checkbox(self, key):
        print("create_dataset_checkbox()")
        entry = self.data_model.get_entry(key)
        if not entry:
            return None
        
        if entry.checkbox is not None:
            return entry.checkbox
        
        file_path, date = key
        name = os.path.basename(file_path)

        # header widget
        header_widget = QWidget()
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_widget.setLayout(header_layout)

        header_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        header_widget.customContextMenuRequested.connect(
            lambda pos, k=key, w=header_widget: self.show_dataset_context_menu(k, w, pos)
        )

        # arrow button
        arrow_btn = QToolButton()
        arrow_btn.setArrowType(Qt.DownArrow)
        arrow_btn.setCheckable(True)
        arrow_btn.setChecked(True)
        # file checkbox
        file_checkbox = QCheckBox(name)
        file_checkbox.setChecked(entry.visible)
        file_checkbox.stateChanged.connect(
            lambda state, e=key: self.on_dataset_toggled(e, state)
        )

        header_layout.addWidget(arrow_btn)
        header_layout.addWidget(file_checkbox)
        header_layout.addStretch()

        self.dataset_layout.addWidget(header_widget)

        # metric container
        metric_container = QWidget()
        metric_layout = QVBoxLayout()
        metric_layout.setContentsMargins(20, 0, 0, 0)
        metric_container.setLayout(metric_layout)

        self.dataset_layout.addWidget(metric_container)

        self.create_metric_checkboxes(entry, metric_layout)

        # arrow toggle operation
        def toggle():
            entry.collapsed = not entry.collapsed

            if entry.collapsed:
                arrow_btn.setArrowType(Qt.RightArrow)
                metric_container.setVisible(False)
            else:
                arrow_btn.setArrowType(Qt.DownArrow)
                metric_container.setVisible(True)

        arrow_btn.clicked.connect(toggle)

        entry.checkbox = file_checkbox # 엔트리에 저장
        entry.header_widget = header_widget
        entry.metric_container = metric_container

        return file_checkbox


    def on_metric_checkbox_changed(self, entry, metric_key, state):
        print("on_metric_checkbox_changed()")
        entry.set_metric_visible(metric_key, bool(state))

        all_checked = all(cb.isChecked() for cb in entry.metric_checkboxes)

        entry.checkbox.blockSignals(True)
        entry.checkbox.setChecked(all_checked)
        entry.checkbox.blockSignals(False)
        if state:
            if not entry.visible:
                entry.visible = True
                entry.checkbox.blockSignals(True)
                entry.checkbox.setChecked(True)
                entry.checkbox.blockSignals(False)
        # entry.visible = all_checked

        self.update_graph()

    def create_metric_checkboxes(self, entry, layout):
        print("create_metric_checkboxes()")
        entry.metric_checkboxes = []

        for metric in ALL_METRICS:

            cb = QCheckBox(metric.label)

            cb.blockSignals(True)
            cb.setChecked(entry.is_metric_visible(metric.key))
            cb.blockSignals(False)

            cb.stateChanged.connect(
                lambda state, e=entry, k=metric.key:
                    self.on_metric_checkbox_changed(e, k, state)
            )

            layout.addWidget(cb)
            entry.metric_checkboxes.append(cb)


    def parse_log_file(self, file_path):
        print("parse_log_file()")
        # Check if it's an Excel file
        if file_path.lower().endswith('.xlsx'):
            self.parse_excel_file(file_path)
            return

        raise Exception("Only Excel files (.xlsx) are supported for this data format.")
        

    def parse_excel_file(self, file_path):
        print("parse_excel_file()")
        """Parse Excel file exported by this application"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Check for Cow Data Format (Check first few rows)
            is_cow_data = False
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
                if row and len(row) > 0:
                    # Check for specific keywords
                    row_str = [str(c) for c in row if c is not None]
                    if "명호" in row_str or "날짜" in row_str:
                        is_cow_data = True
                        break
            
            if is_cow_data:
                self.parse_cow_excel_data(ws)
                wb.close()
                return

            raise Exception("This Excel file does not appear to contain valid Cow Data.")

        except Exception as e:
            raise Exception(f"Failed to parse Excel file: {str(e)}")
        

    def _align_length(self, values, target_len):
        if not values:
            return [None] * target_len
        if len(values) < target_len:
            return values + [None] * (target_len - len(values))
        return values[:target_len]
    

    def _create_curves_for_entry(self, entry):
        # metric_key : (viewbox, color)
        metric_styles = {
            "temps_current":  (self.main_vb, "r"),
            "temps_station": (self.main_vb, "g"),
            "temps_avg":     (self.main_vb, "b"),
            "acts_current":  (self.activity_vb, "c"),
            "acts_station":  (self.activity_vb, "m"),
            "acts_avg":      (self.activity_vb, "y"),
        }

        for metric_key, (vb, color) in metric_styles.items():
            curve = pg.PlotDataItem(
                pen=pg.mkPen(color, width=2),
                clipToView=True
            )

            vb.addItem(curve)
            entry.curves[metric_key] = curve
    

    def parse_cow_excel_data(self, ws):
        print("parse_cow_excel_data()")
        rows = list(ws.iter_rows(values_only=True))
        
        # Find header row with "날짜"
        time_headers = []
        data_start_row = -1
        
        for i, row in enumerate(rows):
            if row and len(row) > 1 and row[1] == "날짜":
                time_headers = row[2:]
                data_start_row = i + 1
                break
        
        if data_start_row == -1:
            return
        
        timestamps = []

        temps_current = []
        temps_station = []
        temps_avg = []
        acts_current = []
        acts_station = []
        acts_avg = []

        last_base_date = None
        # Parse data rows
        for i in range(data_start_row, len(rows)):
            row = rows[i]
            if not row or len(row) < 2: continue
            
            metric = str(row[0]).strip() if row[0] else ""
            date_val = row[1]
            
            # Handle date
            base_date = None
            if isinstance(date_val, datetime):
                base_date = date_val
            elif isinstance(date_val, str):
                try:
                    base_date = datetime.strptime(date_val, '%Y-%m-%d')
                except:
                    pass
            
            if not base_date: continue

            last_base_date = base_date

            # Create timestamps if not done yet
            if not timestamps:
                for t_val in time_headers:
                    if t_val is None:
                        continue
                    dt = None
                    if isinstance(t_val, datetime): # Sometimes time is datetime
                        dt = datetime.combine(base_date.date(), t_val.time())
                    elif hasattr(t_val, 'hour'): # time object
                        dt = datetime.combine(base_date.date(), t_val)
                    elif isinstance(t_val, str):
                        try:
                            h, m = map(int, t_val.split(':'))
                            dt = base_date.replace(hour=h, minute=m)
                        except:
                            pass
                    if dt:
                        timestamps.append(dt)

            # Parse values
            values = []
            for val in row[2:]:
                try:
                    v = float(val)
                    if v == -1: v = None
                except:
                    v = None
                values.append(v)
            
            # Assign to correct list
            n = len(timestamps)

            if metric == "현재 온도":
                temps_current = self._align_length(values, n)
            elif metric == "표본소 온도":
                temps_station = self._align_length(values, n)
            elif metric == "개체 평균 온도":
                temps_avg = self._align_length(values, n)
            elif metric == "현재 활동량":
                acts_current = self._align_length(values, n)
            elif metric == "표본소 활동량":
                acts_station = self._align_length(values, n)
            elif metric == "개체 평균 활동량":
                acts_avg = self._align_length(values, n)

        if last_base_date is None:
            return
        
        file_path = self.file_name   # 임시 식별자
        log_date = last_base_date.date()

        cow = CowData(
            file_path=file_path,
            log_date=log_date,
            timestamps=timestamps,
            temps_current=temps_current,
            temps_station=temps_station,
            temps_avg=temps_avg,
            acts_current=acts_current,
            acts_station=acts_station,
            acts_avg=acts_avg,
        )
        self.data_pool.add(cow)
        entry = DatasetEntry(cow)

        self._create_curves_for_entry(entry)

        self.data_model.add_entry(cow.key, entry)

        axis = self.graph_widget.getAxis('bottom')
        if isinstance(axis, TimeAxisItem):
            axis.set_daily_mode(True)
        self.main_vb.setXRange(0, 86400, padding=0)
        

    def set_excel_visible(self, key, visible):
        entry = self.data_model.set_entries.get(key)
        if not entry:
            return

        entry.visible = visible
        for curve in entry.curves.values():
            curve.setVisible(visible)

    def get_visible_entries(self):
        return [
            entry for entry in self.data_model.get_all_entries()
            if entry.visible
        ]
    
    def clear_plot(self):
        print(">>> CLEAR PLOT")
        for item in list(self.main_vb.addedItems):
            self.main_vb.removeItem(item)
        if hasattr(self, "activity_vb"):
            for item in list(self.activity_vb.addedItems):
                self.activity_vb.removeItem(item)
        for entry in self.data_model.get_all_entries():
            entry.curves.clear()
        # for entry in self.data_model.get_all_entries():
        #     for curve in entry.curves.values():
        #         curve.setVisible(False)

        self.legend.clear()
    

    def _create_crosshair(self):
        self.line_a = pg.InfiniteLine(angle=90, movable=True)
        self.line_b = pg.InfiniteLine(angle=0, movable=False)

        self.main_vb.addItem(self.line_a)
        self.main_vb.addItem(self.line_b)

        self.line_a.sigPositionChanged.connect(self.update_statistics)
        self.line_b.sigPositionChanged.connect(self.update_statistics)


    def _redraw_graph_core(self, visible_entries):
        print("VISIBLE ENTRIES:", len(visible_entries))
        for entry in self.data_model.get_all_entries():
            for curve in entry.curves.values():
                curve.setVisible(False)
        if not visible_entries:
            return

        start_time, time_seconds = self.data_model.get_time_reference()
        self.start_time = start_time
        self.current_time_seconds = time_seconds
        if not time_seconds:
            print("[WARN] No time reference")
            return
        
        self.current_time_seconds = time_seconds

        commands = GraphEngine.prepare_draw_commands(
            visible_entries,
            self.current_time_seconds
        )

        for cmd in commands:
            entry = cmd["entry"]
            metric = cmd["metric"]

            # 🔑 1. curve 없으면 생성
            curve = entry.curves.get(metric)

            if metric.startswith("temps"):
                target_vb = self.main_vb
            else:
                target_vb = self.activity_vb
            
            if curve is None:
                self._create_curves_for_entry(entry)
                curve = entry.curves.get(metric)
                print(f"[CREATE] curve created:", entry.label, metric)
            elif curve.scene() is None:
                target_vb.addItem(curve)
                print(f"[RE-ADD] curve restored:", entry.label, metric)

            # 🔑 2. visible 상태에 따라 처리
            if cmd["x"] and cmd["y"]:
                pairs = [
                    (x, y)
                    for x, y in zip(cmd["x"], cmd["y"])
                    if y is not None and isinstance(y, (int, float))
                ]
                if pairs:
                    xs, ys = zip(*pairs)
                    xs = np.array(xs).flatten()
                    ys = np.array(ys).flatten()
                    curve.setData(xs, ys)
                else:
                    curve.clear()
            else:
                curve.clear()

            # 2. 보이기 / 숨기기는 마지막에만 제어
            curve.setVisible(bool(cmd["visible"]))

        self._create_crosshair()
        self.update_viewbox_geometries()
        self.update_axis_layout()

        # 🔑 OFF → ON 이후 범위 복구 (중요)
        self.main_vb.enableAutoRange(axis=pg.ViewBox.XAxis, enable=True)
        self.main_vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)

        if hasattr(self, "activity_vb"):
            self.activity_vb.enableAutoRange(axis=pg.ViewBox.YAxis, enable=True)

        for cmd in commands:
            print(
                cmd["metric"],
                "visible=", cmd["visible"],
                "x=", len(cmd["x"]) if cmd["x"] else None,
                "y=", len(cmd["y"]) if cmd["y"] else None
            )
        
    
    def update_graph(self):
        print(">>> update_graph ENTER")

        visible_entries = [
            entry for entry in self.data_model.get_all_entries()
            if entry.visible
        ]

        for e in visible_entries:
            print("CURVES:", e.label, e.curves.keys())

        if not visible_entries:
            print("[WARN] No visible entries")
            self.clear_plot()
            return

        # 실제 그리기 위임
        self._redraw_graph_core(visible_entries)

        for entry in self.get_visible_entries():
            print("   AFTER :", entry.visible)

        print("<<< update_graph EXIT")


    def clear_all_datasets(self):
        self.data_pool.clear()

        if hasattr(self, 'dataset_layout'):
            while self.dataset_layout.count():
                item = self.dataset_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()

        self.update_graph()
    

    def update_sensor_graph(self):
        if not self.sensor_data or not self.sensor_data.timestamps:
            return

        self._updating_ranges = True

        # -----------------------------
        # Clear
        # -----------------------------
        self.main_vb.clear()
        self.legend.clear()

        dc = self.sensor_data
        start_time = dc.timestamps[0]
        time_seconds = [(ts - start_time).total_seconds() for ts in dc.timestamps]

        axis = self.graph_widget.getAxis('bottom')
        if isinstance(axis, TimeAxisItem):
            axis.set_start_time(start_time)

        # -----------------------------
        # Helper
        # -----------------------------
        def plot_series(vb, values, pen, name):
            pairs = [(t, v) for t, v in zip(time_seconds, values) if v is not None]
            if not pairs:
                return None
            t, v = zip(*pairs)
            curve = pg.PlotCurveItem(t, v, pen=pen, name=name, clipToView=True)
            vb.addItem(curve)
            return curve

        # -----------------------------
        # Y range 계산
        # -----------------------------
        c_vals = [v for v in dc.c_avg_values if v is not None]
        if c_vals:
            y_min, y_max = self.calculate_y_range_for_quarter(min(c_vals), max(c_vals))
            self.main_vb.setYRange(y_min, y_max)
            self.graph_widget.setLabel('left', 'C Avg (fF)', color='blue')

        # -----------------------------
        # Plot
        # -----------------------------
        self.c_avg_curve = plot_series(
            self.main_vb,
            dc.c_avg_values,
            pg.mkPen('b', width=2),
            'C Avg (fF)'
        )

        self.osc_curve = plot_series(
            self.main_vb,
            dc.osc_values,
            pg.mkPen('teal', width=2),
            'OSC (GHz)'
        )

        # -----------------------------
        # X range
        # -----------------------------
        if time_seconds:
            self.main_vb.setXRange(0, max(time_seconds))

        self.update_viewbox_geometries()
        self.update_axis_layout()

        self._updating_ranges = False

    def update_axis_layout(self):
        """Reorganize axis layout to remove gaps from hidden axes"""
        if not hasattr(self, 'graph_widget'):
            return

        # Set flag to prevent range change callbacks during layout update
        self._updating_ranges = True

        plot_item = self.graph_widget.getPlotItem()
        layout = plot_item.layout

        # Define axes info with their original column positions
        axes_info = [
            (self.activity_axis, 'cow_curr_act_curve', 3)
        ]

        # Remove all axes from layout (safely) - iterate in reverse to avoid index shift issues
        for axis, _, orig_col in axes_info:
            # Find and remove the item from layout (iterate backwards)
            for i in range(layout.count() - 1, -1, -1):
                item = layout.itemAt(i)
                if item == axis:
                    layout.removeAt(i)
                    break

        # Re-add only visible axes, starting from column 3
        col = 3
        for axis, curve_attr, _ in axes_info:
            # Check if axis should be visible (not curve visibility)
            if axis.isVisible():
                layout.addItem(axis, 2, col)
                col += 1

        # Update ViewBox geometries after layout change
        QTimer.singleShot(10, self.update_viewbox_geometries_wrapper)

    def update_viewbox_geometries(self):
        """Update all ViewBox geometries to match the main plot ViewBox"""
        if hasattr(self, 'main_vb') and hasattr(self, 'activity_vb'):
            main_rect = self.main_vb.sceneBoundingRect()
            if hasattr(self, 'activity_vb'):
                self.activity_vb.setGeometry(main_rect)


    def resizeEvent(self, event):
        """Handle window resize events to update ViewBox geometries"""
        super().resizeEvent(event)
        if hasattr(self, 'main_vb') and hasattr(self, 'activity_vb'):
            # Set flag to prevent auto range from being disabled during resize
            self._updating_ranges = True
            # Use a timer to delay the geometry update slightly
            # This ensures the layout has been updated before we query geometries
            QTimer.singleShot(10, self.update_viewbox_geometries_wrapper)

    def update_viewbox_geometries_wrapper(self):
        """Wrapper to update ViewBox geometries and reset the updating flag"""
        try:
            self.update_viewbox_geometries()
        finally:
            # Reset flag after a short delay to ensure all queued signals are processed
            QTimer.singleShot(50, lambda: setattr(self, '_updating_ranges', False))

    def export_to_csv(self, original_file_path):
        # Create CSV filename
        base_name = os.path.splitext(original_file_path)[0]
        csv_path = base_name + '.csv'

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            # Headers in graph order
            writer.writerow(['Time', 'Current Temp', 'Station Temp', 'Avg Temp', 'Current Activity', 'Station Activity', 'Avg Activity'])

            for i, timestamp in enumerate(data.timestamps):
                # Helper to safely get value or empty string
                def get_val(values_list, index):
                    return values_list[index] if index < len(values_list) and values_list[index] is not None else ''

                v1 = get_val(data.cow_current_temp, i)
                v2 = get_val(data.cow_station_temp, i)
                v3 = get_val(data.cow_avg_temp, i)
                v4 = get_val(data.cow_current_activity, i)
                v5 = get_val(data.cow_station_activity, i)
                v6 = get_val(data.cow_avg_activity, i)

                writer.writerow([
                    timestamp.strftime('%Y/%m/%d %H:%M:%S.%f')[:-3],
                    v1, v2, v3, v4, v5, v6
                ])

    def export_to_excel_menu(self):
        """Menu handler for Excel export"""
        if not data.timestamps:
            QMessageBox.warning(self, "Warning", "No data to export. Please open a log file first.")
            return

        # Ask user for save location
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            "",
            "Excel files (*.xlsx);;All files (*)"
        )

        if file_path:
            try:
                self.export_to_excel(file_path)
                # QMessageBox.information(self, "Success", "Data exported to Excel successfully!")
            except PermissionError as e:
                # File is open or no write permission
                QMessageBox.warning(
                    self,
                    "Export Failed",
                    f"Cannot write to file. Please check:\n\n"
                    f"• The file is not currently open in Excel\n"
                    f"• You have write permission for this location\n\n"
                    f"File: {file_path}"
                )
            except Exception as e:
                import traceback
                error_msg = f"Failed to export to Excel:\n\n{str(e)}\n\nFull traceback:\n{traceback.format_exc()}"
                error_dialog = ErrorDialog("Error", error_msg, self)
                error_dialog.exec()

    def export_to_excel(self, original_file_path):
        base_name = os.path.splitext(original_file_path)[0]
        excel_path = base_name + '.xlsx'

        """Export data to Excel file with formatted time column"""
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"

        # Write header row (in graph order)
        headers = ['Time', 'Current Temp', 'Station Temp', 'Avg Temp', 'Current Activity', 'Station Activity', 'Avg Activity']
        ws.append(headers)

        # Make header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write data rows
        for i, timestamp in enumerate(data.timestamps):
             # Helper to safely get value or empty string
            def get_val(values_list, index):
                return values_list[index] if index < len(values_list) and values_list[index] is not None else ''

            v1 = get_val(data.cow_current_temp, i)
            v2 = get_val(data.cow_station_temp, i)
            v3 = get_val(data.cow_avg_temp, i)
            v4 = get_val(data.cow_current_activity, i)
            v5 = get_val(data.cow_station_activity, i)
            v6 = get_val(data.cow_avg_activity, i)

            # Format time as YYYY/MM/DD HH:MM:SS.ms
            time_str = timestamp.strftime('%Y/%m/%d %H:%M:%S') + f'.{timestamp.microsecond // 1000:03d}'

            ws.append([
                time_str,
                v1, v2, v3, v4, v5, v6
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(excel_path)

    def toggle_analysis_lines(self, checked):
        """Toggle visibility of analysis lines A and B"""
        if hasattr(self, 'line_a') and hasattr(self, 'line_b'):
            self.line_a.setVisible(checked)
            self.line_b.setVisible(checked)
            if checked:
                # Position lines at 1/3 and 2/3 of current visible X range
                x_range = self.main_vb.viewRange()[0]  # Get current X-axis range
                x_min, x_max = x_range[0], x_range[1]
                x_span = x_max - x_min

                self.line_a.setPos(x_min + x_span / 4)
                self.line_b.setPos(x_min + 3 * x_span / 4)

                self.update_statistics()
            else:
                # When lines are disabled, show statistics for entire visible range
                self.update_full_range_statistics()

    def on_range_changed(self):
        """Called when the ViewBox range changes (zoom/pan)"""
        # Only update statistics if lines are disabled
        if not self.lines_checkbox.isChecked():
            self.update_full_range_statistics()

    def update_full_range_statistics(self):
        entry = self.data_model.get_primary_entry()
        if not entry:
            return

        cow = entry.cow

        if not cow.timestamps:
            return

        start_time = cow.timestamps[0]
        end_time = cow.timestamps[-1]

        self.stats_time_a.setText(start_time.strftime("%Y/%m/%d %H:%M:%S"))
        self.stats_time_b.setText(end_time.strftime("%Y/%m/%d %H:%M:%S"))

    def update_statistics(self):
        if not hasattr(self, 'line_a') or not hasattr(self, 'line_b'):
            return

        if not self.lines_checkbox.isChecked():
            return

        # Get line positions (in time seconds)
        pos_a = self.line_a.getPos()[0]
        pos_b = self.line_b.getPos()[0]

        # Ensure A is left of B
        if pos_a > pos_b:
            pos_a, pos_b = pos_b, pos_a

        if not self.data_model.set_entries:
            return

        data = next(iter(self.data_model.set_entries.values())).cow

        result = StatisticsEngine.calculate_between_lines(data, pos_a, pos_b)

        # -----------------------------
        # 시간 표시
        # -----------------------------
        if not result or not result["time_range"]:
            self.analysis_text.setText("No data in selected range.")
            return

        time_a, time_b = result["time_range"]

        output = []
        output.append("=== A ↔ B Range Statistics ===\n")
        output.append(f"Time A: {time_a.strftime('%Y/%m/%d %H:%M:%S')}")
        output.append(f"Time B: {time_b.strftime('%Y/%m/%d %H:%M:%S')}\n")

        # -----------------------------
        # 통계 표시
        # -----------------------------
        for key, stats in result["stats"].items():

            avg, min_val, max_val, std_dev = stats

            output.append(f"[{key}]")

            if avg is None:
                output.append("  No data\n")
            else:
                output.append(f"  Avg : {avg:.2f}")
                output.append(f"  Std : {std_dev:.2f}")
                output.append(f"  Min : {min_val:.2f}")
                output.append(f"  Max : {max_val:.2f}\n")

        self.analysis_text.setText("\n".join(output))

    @classmethod
    def _plot_series(cls, vb, time_seconds, values, pen, name):
        values_1d = cls._normalize_1d(values)

        pairs = [(t, v) for t, v in zip(time_seconds, values_1d) if v is not None]
        if not pairs:
            return None

        t, v = zip(*pairs)

        curve = pg.PlotCurveItem(
            np.asarray(t),
            np.asarray(v),
            pen=pen,
            name=name,
            clipToView=True
        )
        vb.addItem(curve)
        return curve
    

    def _draw_temp(self, entry, dc, time_seconds):
        label = dc.label
        curves = []

        configs = [
            ("temp_current", dc.temps_current, pg.mkPen('r', width=2), "Current Temp"),
            ("temp_station", dc.temps_station, pg.mkPen('m', width=2, style=Qt.DashLine), "Station Temp"),
            ("temp_avg", dc.temps_avg, pg.mkPen('orange', width=2), "Avg Temp"),
        ]

        for metric_key, values, pen, name in configs:
            curve = self._plot_series(
                self.main_vb,
                time_seconds,
                values,
                pen,
                f"[{label}] {name}"
            )

            if curve:
                entry.curves[metric_key] = curve



    def _draw_activity(self, entry, dc, time_seconds):
        label = dc.label
        curves = []

        configs = [
            ("act_current", dc.acts_current, pg.mkPen('g', width=2), "Current Activity"),
            ("act_station", dc.acts_station, pg.mkPen('c', width=2, style=Qt.DashLine), "Station Activity"),
            ("act_avg", dc.acts_avg, pg.mkPen('b', width=2), "Avg Activity"),
        ]

        for metric_key, values, pen, name in configs:
            curve = self._plot_series(
                self.activity_vb,
                time_seconds,
                values,
                pen,
                f"[{label}] {name}"
            )

            if curve:
                entry.curves[metric_key] = curve



    def set_x_range(self, time_seconds):
        if time_seconds:
            self.main_vb.setXRange(0, max(time_seconds))
            self.activity_vb.setXLink(self.main_vb)


    def create_analysis_lines(self, time_seconds):
        if not time_seconds:
            return None, None

        max_time = max(time_seconds)

        line_a = pg.InfiniteLine(pos=max_time * 0.25, angle=90, movable=True)
        line_b = pg.InfiniteLine(pos=max_time * 0.75, angle=90, movable=True)

        self.graph_widget.addItem(line_a)
        self.graph_widget.addItem(line_b)

        return line_a, line_b


    def clear_curves(self, entries, main_vb, activity_vb):
        for entry in entries:
            for c in entry.curves["temp"]:
                main_vb.removeItem(c)
            for c in entry.curves["activity"]:
                activity_vb.removeItem(c)

            entry.curves["temp"].clear()
            entry.curves["activity"].clear()


    def remove_dataset(self, key):
        entry = self.data_model.get_entry(key)
        if not entry:
            return

        if hasattr(entry, "header_widget"):
            self.dataset_layout.removeWidget(entry.header_widget)
            entry.header_widget.deleteLater()

        if hasattr(entry, "metric_container"):
            self.dataset_layout.removeWidget(entry.metric_container)
            entry.metric_container.deleteLater()

        entry.visible = False
        self.update_graph()
        self.data_model.remove(key)
          

    def show_dataset_context_menu(self, key, widget, pos):
        menu = QMenu(self)

        delete_action = menu.addAction("삭제")

        global_pos = widget.mapToGlobal(pos)
        action = menu.exec(global_pos)

        if action == delete_action:
            self.remove_dataset(key) 
    
    def apply_y_ranges(self):
        if self.global_temp_range:
            y_min, y_max = self.global_temp_range
            self.main_vb.setYRange(y_min, y_max)

        if self.global_act_range:
            y_min, y_max = self.global_act_range
            self.activity_vb.setYRange(y_min, y_max)
    
    def draw_entries(self, visible_entries, time_seconds):

        self.main_vb.clear()
        self.activity_vb.clear()

        for entry in visible_entries:
            self.draw_entry(entry, time_seconds)

    def draw_entry(self, entry, time_seconds):
        entry.curves = {}

        self._draw_temp(entry, entry.cow, time_seconds)
        self._draw_activity(entry, entry.cow, time_seconds)

    @staticmethod
    def _normalize_1d(values):
        if not values:
            return []

        if isinstance(values, (int, float, np.number)):
            return []

        if isinstance(values, (list, tuple, np.ndarray)):
            if len(values) == 0:
                return []

            if len(values) == 1 and isinstance(values[0], (list, tuple, np.ndarray)):
                return list(values[0])

            return list(values)

        return []